import sys
import os
import time
import openpyxl


def printToXlsx(match_list):
    start_time = time.time()
    try:
        book = openpyxl.load_workbook(r'match_result.xlsx')
    except:
        book = openpyxl.Workbook()
    sheet = book.create_sheet('test', 0)
    i = 1
    for match in match_list:
        sheet.cell(i, 1, match.source_name)
        # sheet.cell(i, 2, match.num)
        sheet.cell(i, 2, match.label)
        sheet.cell(i, 3, match.vul_list)
        # j = 3
        # for patch in match.vul_list:
        #     sheet.cell(i, j, patch)
        #     j += 1
        i += 1
    book.save('match_result.xlsx')
    # print('value of TP: %d' % TP_num)
    # print('value of FP:', FP_num)
    elapsed_time = time.time()-start_time
    print('[+] time cost of printing: %.3f' % elapsed_time)
